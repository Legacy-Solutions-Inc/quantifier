"""
Service for handling export operations
"""

from typing import Dict, List, Any, Optional
from datetime import datetime, timedelta
import uuid
import pandas as pd
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter
import io

class ExportService:
    """Service for handling export operations"""
    
    def __init__(self, db):
        self.db = db
    
    async def export_calculation_results(
        self,
        export_id: str,
        calculation_id: str,
        format: str,
        include_summary: bool,
        include_details: bool,
        user_id: str
    ) -> str:
        """
        Export calculation results in the specified format
        
        Args:
            export_id: Export ID
            calculation_id: Calculation ID
            format: Export format (excel, csv, json)
            include_summary: Whether to include summary statistics
            include_details: Whether to include detailed results
            user_id: User ID
            
        Returns:
            File path of exported file
        """
        try:
            # Get calculation data
            calculation = await self.get_calculation_data(calculation_id)
            
            if not calculation:
                raise Exception("Calculation not found")
            
            # Generate export file
            if format.lower() == "excel":
                file_path = await self.export_to_excel(
                    export_id, calculation, include_summary, include_details
                )
            elif format.lower() == "csv":
                file_path = await self.export_to_csv(
                    export_id, calculation, include_summary, include_details
                )
            elif format.lower() == "json":
                file_path = await self.export_to_json(
                    export_id, calculation, include_summary, include_details
                )
            else:
                raise Exception(f"Unsupported export format: {format}")
            
            # Store export metadata
            await self.store_export_metadata(
                export_id, calculation_id, file_path, format, user_id
            )
            
            return file_path
            
        except Exception as e:
            raise Exception(f"Export failed: {str(e)}")
    
    async def get_calculation_data(self, calculation_id: str) -> Optional[Dict[str, Any]]:
        """Get calculation data from database"""
        try:
            # Get calculation
            calc_response = self.db.table("calculations").select("*").eq("id", calculation_id).execute()
            
            if not calc_response.data:
                return None
            
            calculation = calc_response.data[0]
            
            # Get detailed results
            results_response = self.db.table("calculation_results").select("*").eq("calculation_id", calculation_id).execute()
            
            # Group results by diameter
            results_by_diameter = {}
            for result in results_response.data:
                diameter = result["diameter"]
                if diameter not in results_by_diameter:
                    results_by_diameter[diameter] = []
                
                results_by_diameter[diameter].append({
                    "quantity": result["quantity"],
                    "combination": result["combination"],
                    "lengths": result["lengths"],
                    "combined_length": result["combined_length"],
                    "target": result["target"],
                    "waste": result["waste"],
                    "remaining_pieces": result["remaining_pieces"]
                })
            
            return {
                "calculation": calculation,
                "results": results_by_diameter
            }
            
        except Exception as e:
            raise Exception(f"Failed to get calculation data: {str(e)}")
    
    async def export_to_excel(
        self,
        export_id: str,
        calculation_data: Dict[str, Any],
        include_summary: bool,
        include_details: bool
    ) -> str:
        """Export to Excel format"""
        try:
            wb = Workbook()
            wb.remove(wb.active)
            
            # Define styles
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Create summary sheet if requested
            if include_summary:
                summary_sheet = wb.create_sheet("RSB Summary", 0)
                await self.create_summary_sheet(summary_sheet, calculation_data, header_fill, thin_border)
            
            # Create statistics sheet if requested
            if include_summary:
                stats_sheet = wb.create_sheet("Statistics", 1)
                await self.create_statistics_sheet(stats_sheet, calculation_data, header_fill, thin_border)
            
            # Create detailed results sheets if requested
            if include_details:
                await self.create_detail_sheets(wb, calculation_data, header_fill, thin_border)
            
            # Save to storage
            file_path = f"exports/{export_id}_rsb_results.xlsx"
            
            # Convert workbook to bytes
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            # Upload to Supabase storage
            self.db.storage.from_("exports").upload(file_path, excel_buffer.getvalue())
            
            return file_path
            
        except Exception as e:
            raise Exception(f"Excel export failed: {str(e)}")
    
    async def export_to_csv(
        self,
        export_id: str,
        calculation_data: Dict[str, Any],
        include_summary: bool,
        include_details: bool
    ) -> str:
        """Export to CSV format"""
        try:
            # Create CSV data
            csv_data = []
            
            if include_summary:
                # Add summary data
                for diameter, results in calculation_data["results"].items():
                    for result in results:
                        csv_data.append({
                            "Diameter": diameter,
                            "Quantity": result["quantity"],
                            "Target_Length": result["target"],
                            "Combined_Length": result["combined_length"],
                            "Waste": result["waste"],
                            "Combination": json.dumps(result["combination"]),
                            "Lengths": json.dumps(result["lengths"])
                        })
            
            # Convert to DataFrame
            df = pd.DataFrame(csv_data)
            
            # Save to storage
            file_path = f"exports/{export_id}_rsb_results.csv"
            csv_content = df.to_csv(index=False)
            
            # Upload to Supabase storage
            self.db.storage.from_("exports").upload(file_path, csv_content.encode('utf-8'))
            
            return file_path
            
        except Exception as e:
            raise Exception(f"CSV export failed: {str(e)}")
    
    async def export_to_json(
        self,
        export_id: str,
        calculation_data: Dict[str, Any],
        include_summary: bool,
        include_details: bool
    ) -> str:
        """Export to JSON format"""
        try:
            # Create JSON data
            json_data = {
                "export_info": {
                    "export_id": export_id,
                    "calculation_id": calculation_data["calculation"]["id"],
                    "created_at": datetime.utcnow().isoformat(),
                    "format": "json"
                },
                "calculation": calculation_data["calculation"],
                "results": calculation_data["results"]
            }
            
            # Save to storage
            file_path = f"exports/{export_id}_rsb_results.json"
            json_content = json.dumps(json_data, indent=2, default=str)
            
            # Upload to Supabase storage
            self.db.storage.from_("exports").upload(file_path, json_content.encode('utf-8'))
            
            return file_path
            
        except Exception as e:
            raise Exception(f"JSON export failed: {str(e)}")
    
    async def create_summary_sheet(self, sheet, calculation_data: Dict[str, Any], header_fill, thin_border):
        """Create RSB summary sheet"""
        try:
            # Add title
            sheet.cell(row=1, column=1, value="RSB Summary List")
            sheet.cell(row=1, column=1).font = Font(bold=True, size=12)
            sheet.cell(row=1, column=1).border = thin_border
            
            row = 2
            
            # Add summary entries
            for diameter, results in calculation_data["results"].items():
                for result in results:
                    if result["quantity"] > 0:
                        summary_text = f"{result['quantity']}(pcs) - {diameter}mm(diameter) x {result['target']:.2f}(length) RSB"
                        sheet.cell(row=row, column=1, value=summary_text)
                        sheet.cell(row=row, column=1).border = thin_border
                        row += 1
            
            # Auto-adjust column width
            sheet.column_dimensions[get_column_letter(1)].width = 50
            
        except Exception as e:
            raise Exception(f"Failed to create summary sheet: {str(e)}")
    
    async def create_statistics_sheet(self, sheet, calculation_data: Dict[str, Any], header_fill, thin_border):
        """Create statistics sheet"""
        try:
            # Add title
            sheet.cell(row=1, column=1, value="Statistics Summary")
            sheet.cell(row=1, column=1).font = Font(bold=True, size=12)
            sheet.cell(row=1, column=1).border = thin_border
            
            row = 2
            
            # Calculate and add statistics for each diameter
            for diameter, results in calculation_data["results"].items():
                # Add diameter header
                sheet.cell(row=row, column=1, value=f"Statistics for Diameter {diameter}")
                sheet.cell(row=row, column=1).font = Font(bold=True)
                row += 1
                
                # Calculate statistics
                total_length = sum(r["quantity"] * r["target"] for r in results)
                total_waste = sum(r["waste"] for r in results)
                total_utilized = sum(r["quantity"] * r["combined_length"] for r in results)
                waste_percentage = (total_waste / (total_utilized + total_waste) * 100) if (total_utilized + total_waste) > 0 else 0
                
                # Add statistics
                stats = [
                    ("Total Length (m)", total_length),
                    ("Total Utilized Weight (kg)", total_utilized * (diameter ** 2) / 162),
                    ("Total Commercial Weight (kg)", (total_utilized + total_waste) * (diameter ** 2) / 162),
                    ("Total Waste Weight (kg)", total_waste * (diameter ** 2) / 162),
                    ("Waste Percentage (%)", waste_percentage)
                ]
                
                for label, value in stats:
                    sheet.cell(row=row, column=1, value=label).border = thin_border
                    sheet.cell(row=row, column=2, value=value).border = thin_border
                    row += 1
                
                row += 1  # Add spacing between diameters
            
            # Auto-adjust column widths
            sheet.column_dimensions[get_column_letter(1)].width = 30
            sheet.column_dimensions[get_column_letter(2)].width = 20
            
        except Exception as e:
            raise Exception(f"Failed to create statistics sheet: {str(e)}")
    
    async def create_detail_sheets(self, wb, calculation_data: Dict[str, Any], header_fill, thin_border):
        """Create detailed results sheets for each diameter"""
        try:
            for diameter, results in calculation_data["results"].items():
                # Create sheet for diameter
                sheet_name = f"Diameter {diameter}"
                ws = wb.create_sheet(sheet_name)
                
                # Add headers
                headers = [
                    "Quantity", "Combined Length", "Target Length", "Waste",
                    "Combination", "Lengths", "Remaining Pieces"
                ]
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                
                # Add data
                for row_idx, result in enumerate(results, 2):
                    ws.cell(row=row_idx, column=1, value=result["quantity"]).border = thin_border
                    ws.cell(row=row_idx, column=2, value=result["combined_length"]).border = thin_border
                    ws.cell(row=row_idx, column=3, value=result["target"]).border = thin_border
                    ws.cell(row=row_idx, column=4, value=result["waste"]).border = thin_border
                    ws.cell(row=row_idx, column=5, value=json.dumps(result["combination"])).border = thin_border
                    ws.cell(row=row_idx, column=6, value=json.dumps(result["lengths"])).border = thin_border
                    ws.cell(row=row_idx, column=7, value=json.dumps(result["remaining_pieces"])).border = thin_border
                
                # Auto-fit column widths
                for col in range(1, len(headers) + 1):
                    max_length = 0
                    column = get_column_letter(col)
                    
                    for row in range(1, ws.max_row + 1):
                        cell = ws.cell(row=row, column=col)
                        try:
                            cell_value = str(cell.value) if cell.value is not None else ""
                            max_length = max(max_length, len(cell_value))
                        except:
                            continue
                    
                    adjusted_width = min(max(max_length + 2, 8), 30)
                    ws.column_dimensions[column].width = adjusted_width
                    
        except Exception as e:
            raise Exception(f"Failed to create detail sheets: {str(e)}")
    
    async def store_export_metadata(
        self,
        export_id: str,
        calculation_id: str,
        file_path: str,
        format: str,
        user_id: str
    ) -> None:
        """Store export metadata in database"""
        try:
            export_data = {
                "id": export_id,
                "calculation_id": calculation_id,
                "user_id": user_id,
                "file_path": file_path,
                "filename": f"rsb_results_{calculation_id}.{format}",
                "file_size": 0,  # Will be updated when file is created
                "format": format,
                "created_at": datetime.utcnow(),
                "expires_at": datetime.utcnow() + timedelta(hours=24)
            }
            
            self.db.table("export_history").insert(export_data).execute()
            
        except Exception as e:
            raise Exception(f"Failed to store export metadata: {str(e)}")
    
    async def get_export_file(self, export_id: str, user_id: str) -> Optional[Dict[str, Any]]:
        """Get export file metadata"""
        try:
            response = self.db.table("export_history").select("*").eq("id", export_id).eq("user_id", user_id).execute()
            
            if response.data:
                return response.data[0]
            return None
            
        except Exception as e:
            raise Exception(f"Failed to get export file: {str(e)}")
    
    async def cleanup_expired_exports(self) -> int:
        """Clean up expired export files"""
        try:
            # Get expired exports
            expired_response = self.db.table("export_history").select("*").lt("expires_at", datetime.utcnow()).execute()
            
            expired_count = 0
            for export in expired_response.data:
                try:
                    # Delete file from storage
                    self.db.storage.from_("exports").remove([export["file_path"]])
                    
                    # Delete metadata
                    self.db.table("export_history").delete().eq("id", export["id"]).execute()
                    
                    expired_count += 1
                    
                except Exception as e:
                    print(f"Failed to cleanup export {export['id']}: {e}")
                    continue
            
            return expired_count
            
        except Exception as e:
            raise Exception(f"Failed to cleanup expired exports: {str(e)}")
