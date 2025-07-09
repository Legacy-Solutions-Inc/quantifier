"""
Sidebar component for the RSB Combinator.
Contains control buttons and input fields.
"""

import customtkinter as ctk
from typing import Any, Callable, Dict, Optional
import numpy as np
import pandas as pd

from src.ui.components.drag_drop import DragDropFrame
from src.ui.theme_manager import ThemeManager

<<<<<<< HEAD
=======

>>>>>>> f8f8959bb77f66fc8df23313919a6b0826552b21
class ActionButton(ctk.CTkButton):
    """Custom button with enhanced styling and functionality."""
    
    def __init__(self, parent: Any, text: str, command: Callable, 
                 row: int, tooltip: Optional[str] = None):
        theme = ThemeManager()
        button_style = theme.get_component_style("button")
        
        super().__init__(
            parent,
            text=text,
            command=command,
            font=theme.get_font("bold"),
            corner_radius=button_style["corner_radius"],
            border_width=button_style["border_width"],
            fg_color=theme.get_color("primary", "main"),
            hover_color=theme.get_color("primary", "hover"),
            text_color=theme.get_color("primary", "text")
        )
        self.grid(
            row=row,
            column=0,
            padx=button_style["padding"]["x"],
            pady=button_style["padding"]["y"],
            sticky="ew"
        )
        
        if tooltip:
            self.tooltip = tooltip
            # TODO: Implement tooltip display

class InputField(ctk.CTkFrame):
    """Frame containing a label and input field."""
    
    def __init__(self, parent: Any, label: str, row: int, 
                 default: str = "", validate: Optional[Callable] = None):
        theme = ThemeManager()
        input_style = theme.get_component_style("input")
        
        super().__init__(parent, fg_color="transparent")
        self.grid(
            row=row,
            column=0,
            padx=input_style["padding"]["x"],
            pady=input_style["padding"]["y"],
            sticky="ew"
        )
        
        self.label = ctk.CTkLabel(
            self,
            text=label,
            font=theme.get_font("bold"),
            text_color=theme.get_color("text", "primary")
        )
        self.label.grid(row=0, column=0, padx=5, pady=2, sticky="w")
        
        self.entry = ctk.CTkEntry(
            self,
            font=theme.get_font("regular"),
            border_width=input_style["border_width"],
            corner_radius=input_style["corner_radius"],
            fg_color=theme.get_color("background", "input"),
            text_color=theme.get_color("text", "primary")
        )
        self.entry.grid(row=1, column=0, padx=5, pady=2, sticky="ew")
        self.entry.insert(0, default)
        
        self.validate = validate
        if validate:
            self.entry.bind("<FocusOut>", self._validate)
            
    def get(self) -> str:
        """Get the current value of the input field."""
        return self.entry.get()
        
    def set(self, value: str) -> None:
        """Set the value of the input field."""
        self.entry.delete(0, "end")
        self.entry.insert(0, str(value))
        
    def _validate(self, event: Any) -> None:
        """Validate the input value."""
        if self.validate:
            value = self.get()
            try:
                valid = self.validate(value)
                theme = ThemeManager()
                self.entry.configure(
                    border_color=theme.get_color("validation", "success" if valid else "error")
                )
            except:
                theme = ThemeManager()
                self.entry.configure(
                    border_color=theme.get_color("validation", "error")
                )

class Sidebar(ctk.CTkFrame):
    """Sidebar containing controls and inputs."""
    
    def __init__(self, parent: Any, app: Any):
        theme = ThemeManager()
        super().__init__(
            parent,
            width=150,
            fg_color=theme.get_color("background", "main")
        )
        self.app = app
        
        # Configure grid
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)
        
        # Create button frame
        self.button_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.button_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        
        # Add drag-drop area
        self.drag_drop = DragDropFrame(
            self.button_frame,
            self.app,
            self._handle_file_import
        )
        self.drag_drop.grid(
            row=0,
            column=0,
            sticky="ew",
            padx=10,
            pady=10
        )
        
        # Add diameter selector
        self.diameter_selector = ctk.CTkFrame(self.button_frame, fg_color="transparent")
        self.diameter_selector.grid(row=1, column=0, sticky="ew", padx=10, pady=5)
        
        self.diameter_label = ctk.CTkLabel(
            self.diameter_selector,
            text="Select Diameter:",
            font=theme.get_font("bold"),
            text_color=theme.get_color("text", "primary")
        )
        self.diameter_label.grid(row=0, column=0, padx=5, pady=2, sticky="w")
        
        dropdown_style = theme.get_component_style("dropdown")
        self.diameter_dropdown = ctk.CTkOptionMenu(
            self.diameter_selector,
            font=theme.get_font("regular"),
            fg_color=theme.get_color("background", "dropdown"),
            button_color=theme.get_color("primary", "main"),
            button_hover_color=theme.get_color("primary", "hover"),
            text_color=theme.get_color("text", "primary"),
            dropdown_fg_color=theme.get_color("background", "dropdown"),
            dropdown_text_color=theme.get_color("text", "primary"),
            dropdown_hover_color=theme.get_color("background", "main"),
            command=self._on_diameter_change
        )
        self.diameter_dropdown.grid(row=1, column=0, padx=5, pady=2, sticky="ew")
        self.diameter_dropdown.set("")  # Set empty initial value
        
        # Add input fields
        self.inputs: Dict[str, InputField] = {}
        self._create_inputs()
        
        # Add action buttons
        self._create_buttons()
        
    def _create_inputs(self) -> None:
        """Create input fields."""
        start_row = 2  # Start after drag-drop area and diameter selector
        
        # Target lengths input
        self.inputs["targets"] = InputField(
            self.button_frame,
            "Target Lengths (comma-separated):",
            row=start_row,
<<<<<<< HEAD
            default="12",
=======
            default="12, 10.5, 9, 7.5",
>>>>>>> f8f8959bb77f66fc8df23313919a6b0826552b21
            validate=lambda x: all(float(v) > 0 for v in x.split(","))
        )
        
        # Add stockpile checkbox
        self.stockpile_frame = ctk.CTkFrame(self.button_frame, fg_color="transparent")
        self.stockpile_frame.grid(row=start_row + 1, column=0, sticky="ew", padx=10, pady=5)
        
        theme = ThemeManager()
        self.use_stockpile = ctk.BooleanVar(value=False)
        self.stockpile_checkbox = ctk.CTkCheckBox(
            self.stockpile_frame,
            text="Use Stockpile",
            variable=self.use_stockpile,
            font=theme.get_font("regular"),
            text_color=theme.get_color("text", "primary"),
            checkbox_width=20,
            checkbox_height=20,
            command=self._toggle_stockpile
        )
        self.stockpile_checkbox.grid(row=0, column=0, padx=5, pady=2, sticky="w")
        
    def _create_buttons(self) -> None:
        """Create action buttons."""
        start_row = len(self.inputs) + 3  # +3 for drag-drop area, diameter selector, and stockpile checkbox
        
        # Run combinator button
        self.run_button = ActionButton(
            self.button_frame,
            "Run Combinator",
            self._run_combinator,
            row=start_row,
            tooltip="Generate combinations for current inputs"
        )
        
        # Reset button
        self.reset_button = ActionButton(
            self.button_frame,
            "Reset",
            self._reset_combinator,
            row=start_row + 1,
            tooltip="Reset all inputs and results"
        )
        
        # Export results buttons
        self.export_button = ActionButton(
            self.button_frame,
            "Export Results (RSB Format)",
            self._export_results,
            row=start_row + 2,
            tooltip="Export results with RSB summary and statistics"
        )
        
        self.export_simple_button = ActionButton(
            self.button_frame,
            "Export Results (Simple)",
            self._export_results_simple,
            row=start_row + 3,
            tooltip="Export results as shown in the main window"
        )
        
    def _toggle_stockpile(self) -> None:
        """Handle stockpile checkbox toggle."""
        if self.use_stockpile.get():
            # If stockpile is enabled, check if we have stockpile data
            if not hasattr(self.app.combinator_manager, 'stockpile_data'):
                self.app.show_error("No stockpile data found. Please import a file with a 'STOCKPILE' sheet.")
                self.use_stockpile.set(False)
                return
        # Update display
        self._on_diameter_change(self.diameter_dropdown.get())
        
    def _handle_file_import(self, df: pd.DataFrame) -> None:
        """
        Handle imported file data.
        Args:
            df: DataFrame or dict of DataFrames containing the imported data
        """
        try:
            # If df is a dict (multi-sheet), look for 'LENGTHS' and 'STOCKPILE'
            if isinstance(df, dict):
                available_sheets = list(df.keys())
                if 'LENGTHS' not in df:
                    self.app.show_error(f"Sheet 'LENGTHS' not found. Available sheets: {available_sheets}")
                    return
                lengths_df = df['LENGTHS']
                stockpile_df = df.get('STOCKPILE', None)
            else:
                # Single DataFrame, treat as LENGTHS
                lengths_df = df
                stockpile_df = None
                available_sheets = ['LENGTHS']

            # Set stockpile data if available
            if stockpile_df is not None:
                self.app.combinator_manager.stockpile_data = stockpile_df
                self.stockpile_checkbox.configure(state="normal")
            else:
                self.app.combinator_manager.stockpile_data = None
                self.use_stockpile.set(False)
                self.stockpile_checkbox.configure(state="disabled")

            # Update combinator manager with new data
            self.app.combinator_manager.load_data(lengths_df)

            # Update diameter dropdown
            diameters = self.app.combinator_manager.get_diameters()
            if diameters:
                self.diameter_dropdown.configure(values=[str(d) for d in diameters])
                self.diameter_dropdown.set(str(diameters[0]))
                # Update display for first diameter
                self._on_diameter_change(str(diameters[0]))
            else:
                # Clear dropdown if no diameters
                self.diameter_dropdown.configure(values=[])
                self.diameter_dropdown.set("")

            # Show success message
            self.app.show_success("File imported successfully!")

        except Exception as e:
            self.app.show_error(f"Error processing file data: {str(e)}")
            
    def _on_diameter_change(self, diameter: str) -> None:
        """
        Handle diameter selection change.
        
        Args:
            diameter: Selected diameter as string
        """
        try:
            if not diameter:  # Handle empty selection
                return
                
            # Convert diameter to float and update current diameter
            diameter_float = float(diameter)
            self.app.combinator_manager.set_current_diameter(diameter_float)
            
            # Get current combinator
            combinator = self.app.combinator_manager.get_current_combinator()
            if combinator:
                # Reset to original piece counts
                combinator.pcs = combinator.original_pcs.copy()
                
                # Update display with current results
                df = self.app.create_output_dataframe(cleaned=True)
                self.app.main_window.display_dataframe(df)
                
                # Update waste display
                waste_pct = combinator.get_total_waste_percentage()
                self.app.title_window.show_waste_percentage(waste_pct)
                
        except Exception as e:
            self.app.show_error(f"Error changing diameter: {str(e)}")
            
    def _run_combinator(self) -> None:
        """Run the combinator with current inputs for all diameters."""
        try:
            # Get input values
            targets = [float(x) for x in self.inputs["targets"].get().split(",")]

            # Disable run button after first use
            self.run_button.configure(state="disabled")
            
            # Process each diameter
            for diameter in self.app.combinator_manager.get_diameters():
                # Get combinator for current diameter
                self.app.combinator_manager.set_current_diameter(diameter)
                combinator = self.app.combinator_manager.get_current_combinator()
                
                if not combinator:
                    continue
                    
                # Reset to original piece counts
                combinator.pcs = combinator.original_pcs.copy()
                    
                # Update combinator configuration
                combinator.config.targets = targets
                combinator.targets = np.array(targets)
                
                # If using stockpile, update available pieces
                if self.use_stockpile.get() and self.app.combinator_manager.stockpile_data is not None:
                    stockpile_df = self.app.combinator_manager.stockpile_data
                    # Filter stockpile data for current diameter
                    diameter_stockpile = stockpile_df[stockpile_df['Diameter'] == diameter]
                    # Update available pieces based on stockpile (match both Length and Diameter)
                    for idx, length in enumerate(combinator.lengths):
                        match = diameter_stockpile[diameter_stockpile['Length'] == length]
                        if not match.empty:
                            stock_qty = int(match.iloc[0]['Quantity'])
                            combinator.pcs[idx] = min(combinator.original_pcs[idx], stock_qty)
                        else:
                            combinator.pcs[idx] = combinator.original_pcs[idx]
                
                # Run combination process
                combinator.iterate_combinations()
                
                # Calculate waste
                combinator.calculate_waste()
            
            # Display results for current diameter
            df = self.app.create_output_dataframe(cleaned=True)
            self.app.main_window.display_dataframe(df)
            
            # Update waste display
            current_combinator = self.app.combinator_manager.get_current_combinator()
            if current_combinator:
                waste_pct = current_combinator.get_total_waste_percentage()
                self.app.title_window.show_waste_percentage(waste_pct)
            
        except Exception as e:
            self.app.show_error(f"Error running combinator: {str(e)}")
            
    def _reset_combinator(self) -> None:
        """Reset the combinator and clear display."""
        self.app.combinator_manager.reset()
        self.diameter_dropdown.configure(values=[])
        self.diameter_dropdown.set("")
        self.app.main_window.clear_table()
        self.app.title_window.clear_details()
        # Re-enable run button
        self.run_button.configure(state="normal")
        
    def _export_results(self) -> None:
        """Export current results to Excel with RSB summary and statistics."""
        try:
            from tkinter import filedialog
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
            from openpyxl.utils import get_column_letter
            
            # Ask for save location
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Save Results"
            )
            
            if not file_path:
                return
                
            # Create workbook
            wb = Workbook()
            wb.remove(wb.active)
            
            # Define styles
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            number_format = '#,##0.00'
            
            # Create RSB Summary sheet
            summary_sheet = wb.create_sheet("RSB Summary", 0)
            summary_sheet.cell(row=1, column=1, value="RSB Summary List")
            summary_sheet.cell(row=1, column=1).font = Font(bold=True, size=12)
            summary_sheet.cell(row=1, column=1).border = thin_border
            
            # Create Statistics sheet
            stats_sheet = wb.create_sheet("Statistics", 1)
            stats_sheet.cell(row=1, column=1, value="Statistics Summary")
            stats_sheet.cell(row=1, column=1).font = Font(bold=True, size=12)
            stats_sheet.cell(row=1, column=1).border = thin_border
            
            # Process each diameter
            row = 2  # Start row for RSB summary
            stats_row = 2  # Start row for statistics
            grand_total = {
                'total_length': 0,
                'total_weight': 0,
                'total_commercial_weight': 0,
                'total_waste_weight': 0,
                'total_waste_percentage': 0
            }
            
            for diameter, combinator in sorted(self.app.combinator_manager.combinators.items()):
                # Group results by target length
                target_groups = {}
                for result in combinator.results:
                    if result.target not in target_groups:
                        target_groups[result.target] = 0
                    target_groups[result.target] += result.quantity
                
                # Calculate statistics for current diameter
                diameter_stats = {
                    'total_length': 0,
                    'total_weight': 0,
                    'total_commercial_weight': 0,
                    'total_waste_weight': 0
                }
                
                # Add RSB summary entries
                for target, quantity in sorted(target_groups.items(), reverse=True):
                    if quantity > 0:
                        formatted_summary = f"{quantity}(pcs) - {diameter}mm(diameter) x {target:.2f}(length) RSB"
                        cell = summary_sheet.cell(row=row, column=1, value=formatted_summary)
                        cell.border = thin_border
                        diameter_stats['total_length'] += quantity * target
                        row += 1
                
                # Calculate weights
                diameter_stats['total_commercial_weight'] = diameter_stats['total_length'] * (diameter ** 2) / 162
                waste_percentage = combinator.get_total_waste_percentage()
                diameter_stats['total_waste_weight'] = (diameter_stats['total_commercial_weight'] * waste_percentage / 100)
                
                diameter_stats['total_weight'] = (diameter_stats['total_commercial_weight'] * (1 - (waste_percentage / 100)))
                
                # Add diameter statistics
                stats_sheet.cell(row=stats_row, column=1, value=f"Statistics for Diameter {diameter}")
                stats_sheet.cell(row=stats_row, column=1).font = Font(bold=True)
                stats_row += 1
                
                stats = [
                    ("Total Length (m)", diameter_stats['total_length']),
                    ("Total Utilized Weight (kg)", diameter_stats['total_weight']),
                    ("Total Commercial Weight (kg)", diameter_stats['total_commercial_weight']),
                    ("Total Waste Weight (kg)", diameter_stats['total_waste_weight']),
                    ("Waste Percentage (%)", waste_percentage)
                ]
                
                for label, value in stats:
                    cell_label = stats_sheet.cell(row=stats_row, column=1, value=label)
                    cell_value = stats_sheet.cell(row=stats_row, column=2, value=value)
                    cell_label.border = thin_border
                    cell_value.border = thin_border
                    if "Weight" in label or label.startswith("Total Length"):
                        cell_value.number_format = number_format
                    elif "Percentage" in label:
                        cell_value.number_format = '0.00"%"'
                    stats_row += 1
                
                # Update grand totals
                grand_total['total_length'] += diameter_stats['total_length']
                grand_total['total_weight'] += diameter_stats['total_weight']
                grand_total['total_commercial_weight'] += diameter_stats['total_commercial_weight']
                grand_total['total_waste_weight'] += diameter_stats['total_waste_weight']
                stats_row += 1
            
            # Add grand total statistics
            stats_sheet.cell(row=stats_row, column=1, value="Grand Total Statistics")
            stats_sheet.cell(row=stats_row, column=1).font = Font(bold=True, size=12)
            stats_row += 1
            
            if grand_total['total_commercial_weight'] > 0:
                grand_total['total_waste_percentage'] = (
                    grand_total['total_waste_weight'] / grand_total['total_commercial_weight'] * 100
                )
            
            grand_stats = [
                ("Total Length (m)", grand_total['total_length']),
                ("Total Utilized Weight (kg)", grand_total['total_weight']),
                ("Total Commercial Weight (kg)", grand_total['total_commercial_weight']),
                ("Total Waste Weight (kg)", grand_total['total_waste_weight']),
                ("Overall Waste Percentage (%)", grand_total['total_waste_percentage'])
            ]
            
            for label, value in grand_stats:
                cell_label = stats_sheet.cell(row=stats_row, column=1, value=label)
                cell_value = stats_sheet.cell(row=stats_row, column=2, value=value)
                cell_label.border = thin_border
                cell_value.border = thin_border
                if "Weight" in label or label.startswith("Total Length"):
                    cell_value.number_format = number_format
                elif "Percentage" in label:
                    cell_value.number_format = '0.00"%"'
                stats_row += 1
            
            # Auto-adjust column widths
            summary_sheet.column_dimensions[get_column_letter(1)].width = 50  # RSB Summary column
            stats_sheet.column_dimensions[get_column_letter(1)].width = 30  # Labels column
            stats_sheet.column_dimensions[get_column_letter(2)].width = 20  # Values column
            
            # Save the workbook
            wb.save(file_path)
            self.app.show_success("Results exported successfully!")
            
        except Exception as e:
            self.app.show_error(f"Error exporting results: {str(e)}")
            
    def _export_results_simple(self) -> None:
        """Export current results to Excel using simple format from main window."""
        try:
            from tkinter import filedialog
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
            from openpyxl.utils import get_column_letter
            
            # Ask for save location
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Save Results"
            )
            
            if not file_path:
                return
                
            # Create workbook
            wb = Workbook()
            wb.remove(wb.active)
            
            # Define styles
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Process each diameter
            for diameter, combinator in sorted(self.app.combinator_manager.combinators.items()):
                # Set current diameter to get the right data
                self.app.combinator_manager.set_current_diameter(diameter)
                df = self.app.create_output_dataframe(cleaned=True)
                
                # Replace zeros with "-" in the DataFrame
                df = df.replace(0, "-")
                
                if df.empty:
                    continue
                    
                # Create sheet for current diameter
                sheet_name = f"Diameter {diameter}"
                ws = wb.create_sheet(sheet_name)
                
                # Write headers
                for col, header in enumerate(df.columns, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    
                # Write data
                for row_idx, row in enumerate(df.values, 2):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center')
                        
                        # Handle numeric values and formatting
                        if pd.notnull(value):  # Check if value is not NaN
                            if value == "-":  # Keep the dash as is
                                cell.value = "-"
                            elif isinstance(value, (int, float)):
                                cell.value = value
                                cell.number_format = "0.00"  # Use decimal format
                            else:
                                cell.value = value
                        else:
                            cell.value = ""  # Empty cell for NaN values
                
                # Auto-fit column widths
                for col in range(1, len(df.columns) + 1):
                    max_length = 0
                    column = get_column_letter(col)
                    
                    # Iterate through cells to find max width needed
                    for row in range(1, ws.max_row + 1):
                        cell = ws.cell(row=row, column=col)
                        try:
                            cell_value = str(cell.value) if cell.value is not None else ""
                            max_length = max(max_length, len(cell_value))
                        except:
                            continue
                    
                    # Set width with padding (min 8, max 30)
                    adjusted_width = min(max(max_length + 2, 8), 30)
                    ws.column_dimensions[column].width = adjusted_width
                    
            # Save the workbook
            wb.save(file_path)
            self.app.show_success("Results exported successfully!")
            
        except Exception as e:
            self.app.show_error(f"Error exporting results: {str(e)}")